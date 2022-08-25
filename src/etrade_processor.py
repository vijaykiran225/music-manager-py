import csv
from openpyxl import load_workbook
import json
import babel.numbers

debug= False
g_today_cost=106.51
g_today_usd_inr= 75.795
g_xlsx_src="/Users/vj/Downloads/ByBenefitType.xlsx"


def getIndices(data):
    indices=[]
    for i in range(1,data.max_column+1):
        indices.append(data.cell(1,i).value)
    return indices

def formatData(data):
    indices=getIndices(data)

    d=[]
    for row in data:
        i=0
        x=dict()
        for col in row:
            aKey = indices[i]
            if aKey not in x and col.value != None:
                x[aKey]= col.value
            i=i+1
        d.append(x)
    d.pop(0)
    return d

def espp(data):
    if debug:
        print("processing espp")
    d=formatData(data)
    if debug:
        print(json.dumps(d,indent=2))
    
    total_shares=0
    ans=0
    for entry in d:
        if entry["Record Type"] == "Purchase":
            price=(entry["Purchase Price"])
            qty=(entry["Sellable Qty."])
            if debug:
                print(price,qty)
            if price and qty:
                ans = ans + (price *qty)
                total_shares=total_shares+qty

    if debug:
        print("done espp")
    return {
        "buying_price":ans,
        "total_shares":total_shares,
        "type":"ESPP"
    }

def performance_shares(data):
    if debug:
        print("processing performance_shares")
    d=formatData(data)
    if debug:
        print(json.dumps(d,indent=2))
    
    total_shares=0
    ans=0
    for entry in d:
        if entry["Record Type"] == "Sellable Shares":
            
            price=(entry["Est. Cost Basis (per share):"])
            qty=(entry["Sellable Qty."])
            if debug:
                print(price,qty)
            if price and qty:
                ans = ans + (price *qty)
                total_shares=total_shares+qty

    if debug:
        print("done performance_shares")
    return {
        "buying_price":ans,
        "total_shares":total_shares,
        "type":"performance_shares"
    }

def restricted_stock(data):
    if debug:
        print("processing restricted_stock")
    d=formatData(data)
    if debug:
        print(json.dumps(d,indent=2))

    total_shares=0
    ans=0
    for entry in d:
        if entry["Record Type"] == "Sellable Shares":
            
            price=(entry["Est. Cost Basis (per share):"])
            qty=(entry["Sellable Qty."])
            if debug:
                print(price,qty)
            if price and qty:
                ans = ans + (price *qty)
                total_shares=total_shares+qty

    if debug:
        print("done restricted_stock")
    return {
        "buying_price":ans,
        "total_shares":total_shares,
        "type":"restricted_stock"
    }

def other_holdings(data):
    if debug:
        print("processing other_holdings")
    # d=formatData(data)
    # print(json.dumps(d,indent=2))
    if debug:
        print("done other_holdings")
    return {
        "buying_price":0,
        "total_shares":0,
        "type":"other_holdings"
    }

handlers = {
    'ESPP' : espp,
    'Performance Shares':performance_shares,
    'Restricted Stock':restricted_stock,
    'Other Holdings':other_holdings,
}

def processResponses(data,today_cost,today_usd_conv_rate):
    sellableShares=0
    sellingTotalVolume=0
    buyingTotalVolume=0
    gain=0
    for datum in data:
        sellableShares = sellableShares + datum["total_shares"] 
        selling_price= today_cost *datum["total_shares"] *today_usd_conv_rate
        cost_price=datum["buying_price"] *today_usd_conv_rate
        print(datum["type"],selling_price ,cost_price)
        gain=gain+selling_price -cost_price
        sellingTotalVolume = sellingTotalVolume + selling_price
        buyingTotalVolume=buyingTotalVolume+cost_price

    

    print("count  ==> ",babel.numbers.format_decimal(sellableShares,  locale='en_IN'))
    print("se_vol ==> ",babel.numbers.format_currency(sellingTotalVolume, 'INR', locale='en_IN'))
    print("bu_vol ==> ",babel.numbers.format_currency(buyingTotalVolume, 'INR', locale='en_IN'))
    print("gain   ==> ",babel.numbers.format_currency(gain, 'INR', locale='en_IN'))
    print("stock  ==> ",babel.numbers.format_currency(today_cost, 'USD', locale='en_US'))
    print("usd_c  ==> ",babel.numbers.format_number(today_usd_conv_rate, locale='en_IN'))
    pass

def readFromXLSX(fileName):    
    workBook = load_workbook(filename = fileName)
    # print(workBook.sheetnames)
    responses = [] 
    for sheet_name in workBook.sheetnames:
        if sheet_name in handlers:
            logic= handlers[sheet_name]
            data= workBook[sheet_name]

            responses.append(logic(data))
            # print(response)
        else:
            print("dont know this one")
    return responses

def readFromCSV(fileName):
    reader= csv.DictReader(open(fileName))

    for r in reader:
        print(r)
    pass

def main():
    csv_src="/Users/vj/Downloads/espp_csv.csv"
    responses=readFromXLSX(g_xlsx_src)
    processResponses(responses,g_today_cost,g_today_usd_inr)
    print("done")

main()
